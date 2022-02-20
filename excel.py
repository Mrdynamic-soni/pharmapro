import webbrowser
import openpyxl

path = "C:\\Users\\Acer\\Desktop\\PHARMAPRO\\Data.xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# method to print all the colmns from excel 
for i in range(1,sheet_obj.max_column+1):
    print(sheet_obj.cell(row=1,column=i).value)

# append data on the file
sheet_obj.cell(row=2,column=2).value = "soni"
wb_obj.save(path)